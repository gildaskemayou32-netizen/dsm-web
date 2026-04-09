"""Export Excel/CSV"""
import io, csv
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _side(c="1e2645"): return Side(style="thin", color=c)
def _brd(): s=_side(); return Border(left=s,right=s,top=s,bottom=s)
def _fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
def _font(bold=False,color="dde3f5",size=10):
    return Font(bold=bold,color=color.lstrip("#"),size=size,name="Segoe UI")
def _al(h="left",v="center"):
    return Alignment(horizontal=h,vertical=v)


def export_excel(transactions: list) -> bytes:
    wb = openpyxl.Workbook()
    _sheet_tx(wb, transactions)
    _sheet_recap(wb, transactions)
    _sheet_clients(wb, transactions)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_tx(wb, txs):
    ws = wb.active; ws.title = "Transactions"
    ws.sheet_view.showGridLines = False
    headers = ["#","Client","Reçu (FCFA)","Transport","Autres",
               "Bénéfice (FCFA)","Marge %","Statut","Notes","Date"]
    widths  = [5,26,16,14,12,18,10,10,22,13]
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(row=1,column=c,value=h)
        cell.fill=_fill("0b1020"); cell.font=_font(True,"f0c040")
        cell.alignment=_al("center"); cell.border=_brd()
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[1].height=22
    for ri,t in enumerate(txs,2):
        bg="0f1424" if ri%2==0 else "131829"
        ben=t.benefice; bc="10e080" if ben>=0 else "f0504a"
        vals=[ri-1,t.client,t.montant_recu,t.transport,t.autres,
              ben,f"{t.marge:.1f}%",t.statut,t.notes or "",t.date]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=ri,column=c,value=v)
            cell.fill=_fill(bg); cell.border=_brd()
            cell.font=_font(c==6,bc if c==6 else "dde3f5")
            cell.alignment=_al("right" if c in(3,4,5,6,7) else "center" if c in(1,8,10) else "left")
            if c in(3,4,5,6): cell.number_format="#,##0"
    tr=len(txs)+2
    ws.cell(row=tr,column=2,value="TOTAL GÉNÉRAL").font=_font(True,"f0c040",11)
    for col,attr in [(3,"montant_recu"),(4,"transport"),(5,"autres")]:
        v=sum(getattr(t,attr) for t in txs)
        cell=ws.cell(row=tr,column=col,value=v)
        cell.number_format="#,##0"; cell.font=_font(True,"fde68a",11)
        cell.alignment=_al("right")
    bt=sum(t.benefice for t in txs)
    c=ws.cell(row=tr,column=6,value=bt)
    c.number_format="#,##0"
    c.font=_font(True,"10e080" if bt>=0 else "f0504a",11)
    c.alignment=_al("right")
    for col in range(1,11):
        ws.cell(row=tr,column=col).fill=_fill("0b2040")
        ws.cell(row=tr,column=col).border=_brd()
    ws.freeze_panes="A2"


def _sheet_recap(wb, txs):
    ws=wb.create_sheet("Récapitulatif")
    ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=34
    ws.column_dimensions["B"].width=22
    total_recu=sum(t.montant_recu for t in txs)
    total_frais=sum(t.total_frais for t in txs)
    total_ben=sum(t.benefice for t in txs)
    nb=len(txs)
    rows=[("Métrique","Valeur",True),
          ("Total revenus reçus",total_recu,False),
          ("Total frais",total_frais,False),
          ("Bénéfice net total",total_ben,False),
          ("Nombre de transactions",nb,False),
          ("Bénéfice moyen/transaction",round(total_ben/nb,2) if nb else 0,False),
          ("Marge moyenne",f"{round(total_ben/total_recu*100,1) if total_recu else 0}%",False)]
    for ri,(lab,val,head) in enumerate(rows,1):
        for c,v in enumerate([lab,val],1):
            cell=ws.cell(row=ri,column=c,value=v)
            cell.border=_brd(); cell.alignment=_al("left" if c==1 else "right")
            if head:
                cell.fill=_fill("0b1020"); cell.font=_font(True,"f0c040")
            else:
                cell.fill=_fill("0f1424" if ri%2==0 else "131829")
                cell.font=_font()
                if c==2 and isinstance(val,(int,float)): cell.number_format="#,##0"


def _sheet_clients(wb, txs):
    ws=wb.create_sheet("Par Client")
    ws.sheet_view.showGridLines=False
    for i,w in enumerate([26,14,16,18,10,13],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    data=defaultdict(lambda:{"tx":0,"recu":0.0,"ben":0.0})
    for t in txs:
        data[t.client]["tx"]+=1; data[t.client]["recu"]+=t.montant_recu
        data[t.client]["ben"]+=t.benefice
    headers=["Client","Transactions","Reçu (FCFA)","Bénéfice (FCFA)","Marge %","Dernière tx"]
    for c,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=c,value=h)
        cell.fill=_fill("0b1020"); cell.font=_font(True,"f0c040")
        cell.border=_brd(); cell.alignment=_al("center")
    sorted_c=sorted(data.items(),key=lambda x:x[1]["ben"],reverse=True)
    for ri,(client,d) in enumerate(sorted_c,2):
        bg="0f1424" if ri%2==0 else "131829"
        marge=round(d["ben"]/d["recu"]*100,1) if d["recu"] else 0
        last=max((t.date for t in txs if t.client==client),default="—")
        vals=[client,d["tx"],d["recu"],d["ben"],f"{marge:.1f}%",last]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row=ri,column=c,value=v)
            cell.fill=_fill(bg); cell.border=_brd()
            cell.font=_font(c==4,"10e080" if c==4 and d["ben"]>=0 else "f0504a" if c==4 else "dde3f5")
            cell.alignment=_al("left" if c==1 else "right" if c in(2,3,4,5) else "center")
            if c in(3,4): cell.number_format="#,##0"


def export_csv(transactions: list) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID","Client","Reçu","Transport","Autres",
                     "Bénéfice","Marge%","Statut","Date"])
    for t in transactions:
        writer.writerow([t.id, t.client, t.montant_recu, t.transport,
                         t.autres, t.benefice, t.marge, t.statut, t.date])
    return output.getvalue()
